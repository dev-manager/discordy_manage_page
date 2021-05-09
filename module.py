import hashlib
import re
import openpyxl


class auth:
    def __init__(self):
        self.account = {}
        self.logined = {}
        wb = openpyxl.load_workbook('account.xlsx')
        sheet = wb[wb.sheetnames[0]]
        for i in range(2, sheet.max_row + 1):
            self.account[sheet['A' + str(i)].value] = sheet['B' + str(i)].value
            self.logined[sheet['A' + str(i)].value] = False
            
    
    def password(self, pwd:str):
        pattern = re.compile(r'([a-zA-Z0-9_]{8,20})([a-zA-Z0-9_])')
        if len(pattern.match(pwd).groups()) >= 2:
            return None
        else:
            return hashlib.sha256(pwd.encode())
    
    
    def authenticate(self, Id:str, pwd:str):
            try:
                matched = self.account[Id] == pwd
                if matched:
                    return True
                elif not matched:
                    return "Password doesn\'t match with Id"
            except KeyError:
                return "Id doesn\'t found"


if __name__ == '__main__':
    a = auth()
    print(a.account)
    print(a.authenticate(Id='geon0601', pwd='8c6976e5b5410415bde908bd4dee15dfb167a9c873fc4bb8a81f6f2ab448a918'))
